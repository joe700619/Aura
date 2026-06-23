"""
產生「服務費用」與「公費分攤」匯入範本 .xlsx。
一次性工具，跑完即可：python tools/gen_import_templates.py
輸出到 docs/import_templates/。
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'docs', 'import_templates')
os.makedirs(OUT_DIR, exist_ok=True)

HEADER_FILL = PatternFill('solid', fgColor='4F81BD')
HEADER_FONT = Font(bold=True, color='FFFFFF')
CENTER = Alignment(horizontal='center', vertical='center')


def style_header(ws, headers, widths):
    for c, (title, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        ws.column_dimensions[cell.column_letter].width = width
    ws.freeze_panes = 'A2'


def set_text_columns(ws, text_cols, n_rows=200):
    """把指定欄位設文字格式，避免統編/員工編號前導 0 掉字。
    必須在 append 完範例列後才呼叫，否則會撐高 max_row 把資料擠到後面。"""
    for c in text_cols:
        letter = ws.cell(row=1, column=c).column_letter
        for r in range(2, n_rows + 2):
            ws[f'{letter}{r}'].number_format = '@'


# ── 範本 1：服務費用 ───────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = '服務費用'
headers = ['統一編號', '服務費用', '帳簿費用', '收費月份', '收費週期', '生效日', '結束日']
widths = [16, 12, 12, 12, 22, 14, 14]
style_header(ws, headers, widths)

ws.cell(row=1, column=5).comment = Comment(
    '可填中文或代碼：\n'
    '月繳 / 雙月 / 半年 / 年繳 /\n'
    '雙月(自動扣款) / 半年(預繳收費)', 'template', height=120, width=220)
ws.cell(row=1, column=4).comment = Comment(
    '無特定收費月份可填 13', 'template')
ws.cell(row=1, column=1).comment = Comment(
    '對應已匯入的記帳客戶；一客戶一筆', 'template')

examples = [
    ['12345678', 3000, 500, 13, '雙月', '2026-01-01', ''],
    ['98765432', 2500, 0, 1, '月繳', '2026-01-01', ''],
]
for row in examples:
    ws.append(row)
set_text_columns(ws, text_cols=(1,))

wb.save(os.path.join(OUT_DIR, '服務費用_匯入範本.xlsx'))


# ── 範本 2：公費分攤 ───────────────────────────────────────────
wb2 = Workbook()
ws2 = wb2.active
ws2.title = '公費分攤'
headers2 = ['統一編號', '員工編號', '執行項目', '比例']
widths2 = [16, 14, 36, 10]
style_header(ws2, headers2, widths2)

ws2.cell(row=1, column=1).comment = Comment(
    '同一統編可有多列（多個分攤人），\n會自動合併成一筆', 'template', height=80, width=220)
ws2.cell(row=1, column=2).comment = Comment(
    '用員工編號反查；請保持文字格式\n避免前導 0 掉字 (001→1)', 'template', height=80, width=220)
ws2.cell(row=1, column=4).comment = Comment(
    '百分比數字，例如 60 代表 60%', 'template')

examples2 = [
    ['12345678', 'E001', '記帳作業', 60],
    ['12345678', 'E002', '稅務申報', 40],
    ['98765432', 'E001', '記帳作業', 100],
]
for row in examples2:
    ws2.append(row)
set_text_columns(ws2, text_cols=(1, 2))

wb2.save(os.path.join(OUT_DIR, '公費分攤_匯入範本.xlsx'))

print('已產生：')
print(' -', os.path.abspath(os.path.join(OUT_DIR, '服務費用_匯入範本.xlsx')))
print(' -', os.path.abspath(os.path.join(OUT_DIR, '公費分攤_匯入範本.xlsx')))
