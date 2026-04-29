from decimal import Decimal, InvalidOperation
from collections import OrderedDict

from django.contrib import admin
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import path
from import_export.admin import ImportExportModelAdmin
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .models import Account, Voucher, VoucherDetail, Receivable, FixedAsset, Collection, ReceivableNotification
from .models.pre_collection import PreCollection

@admin.register(Account)
class AccountAdmin(ImportExportModelAdmin):
    list_display = ('code', 'name', 'category', 'is_active')
    list_filter = ('category', 'is_active')
    search_fields = ('code', 'name')
    ordering = ('code',)
    list_per_page = 25

class VoucherDetailInline(admin.TabularInline):
    model = VoucherDetail
    extra = 2
    autocomplete_fields = ('account',)

_VOUCHER_HEADERS = [
    '傳票編號', '傳票日期(YYYY-MM-DD)', '摘要', '狀態(草稿/已過帳)',
    '科目代碼', '借方金額', '貸方金額',
    '往來對象/統編', '部門', '專案', '備註',
]

_STATUS_MAP = {
    '草稿': Voucher.Status.DRAFT, 'DRAFT': Voucher.Status.DRAFT,
    '已過帳': Voucher.Status.POSTED, 'POSTED': Voucher.Status.POSTED,
}


def _parse_amount(val) -> Decimal:
    try:
        return Decimal(str(val or 0)).quantize(Decimal('1'))
    except InvalidOperation:
        return Decimal(0)


def _parse_date(val):
    from datetime import datetime
    if hasattr(val, 'date'):
        return val.date()
    try:
        return datetime.strptime(str(val).strip(), '%Y-%m-%d').date()
    except ValueError:
        return None


@admin.register(Voucher)
class VoucherAdmin(admin.ModelAdmin):
    list_display = ('voucher_no', 'date', 'status', 'description', 'created_by', 'is_deleted')
    list_filter = ('status', 'is_deleted', 'date')
    search_fields = ('voucher_no', 'description')
    inlines = [VoucherDetailInline]
    list_per_page = 25
    actions = ['restore_vouchers']
    change_list_template = 'admin/internal_accounting/voucher/change_list.html'

    def get_queryset(self, _request):
        return self.model._default_manager.all()

    @admin.action(description='還原選取的傳票（取消軟刪除）')
    def restore_vouchers(self, request, queryset):
        updated = queryset.update(is_deleted=False)
        self.message_user(request, f'已成功還原 {updated} 筆傳票。')

    def save_model(self, request, obj, form, change):
        if not obj.pk:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import/', self.admin_site.admin_view(self.import_view), name='voucher_import'),
            path('import/template/', self.admin_site.admin_view(self.download_template), name='voucher_template'),
            path('export/', self.admin_site.admin_view(self.export_view), name='voucher_export'),
        ]
        return custom_urls + urls

    # ------------------------------------------------------------------ import

    def import_view(self, request):
        if request.method == 'POST':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                self.message_user(request, '請選擇 Excel 檔案', level='error')
                return redirect('.')
            try:
                result = self._do_import(request, excel_file)
                if result['errors']:
                    for err in result['errors']:
                        self.message_user(request, err, level='error')
                if result['created']:
                    self.message_user(
                        request,
                        f"成功匯入 {result['created']} 張傳票，共 {result['details']} 筆分錄",
                    )
                return redirect('../')
            except Exception as e:
                self.message_user(request, f'匯入失敗：{e}', level='error')
                return redirect('.')

        context = self.admin_site.each_context(request)
        context['title'] = '匯入傳票資料'
        context['opts'] = self.model._meta
        context['headers'] = _VOUCHER_HEADERS
        return render(request, 'admin/internal_accounting/voucher/import.html', context)

    def _do_import(self, request, excel_file):
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(r)]

        # Pre-load accounts
        account_map = {a.code: a for a in Account.objects.filter(is_active=True)}

        # Group rows into vouchers using an OrderedDict to preserve order.
        # Key: voucher_no if provided, else (date_str, description)
        groups: OrderedDict = OrderedDict()
        for i, row in enumerate(rows, start=2):
            vno = str(row[0] or '').strip()
            date_str = str(row[1] or '').strip()
            desc = str(row[2] or '').strip()
            key = vno if vno else f'__row_{i}__{date_str}__{desc}'
            if key not in groups:
                groups[key] = []
            groups[key].append((i, row))

        errors = []
        created_count = 0
        detail_count = 0

        with transaction.atomic():
            for key, row_pairs in groups.items():
                first_row = row_pairs[0][1]
                date = _parse_date(first_row[1])
                if not date:
                    errors.append(f'第 {row_pairs[0][0]} 列：日期格式錯誤（{first_row[1]}），應為 YYYY-MM-DD')
                    continue

                details_data = []
                row_errors = []
                for row_no, row in row_pairs:
                    code = str(row[4] or '').strip()
                    if not code:
                        row_errors.append(f'第 {row_no} 列：科目代碼不能為空')
                        continue
                    account = account_map.get(code)
                    if not account:
                        row_errors.append(f'第 {row_no} 列：科目代碼 "{code}" 不存在')
                        continue
                    details_data.append({
                        'account': account,
                        'debit': _parse_amount(row[5]),
                        'credit': _parse_amount(row[6]),
                        'company_id': str(row[7] or '').strip(),
                        'department': str(row[8] or '').strip(),
                        'project': str(row[9] or '').strip(),
                        'remark': str(row[10] or '').strip(),
                    })

                if row_errors:
                    errors.extend(row_errors)
                    continue

                total_debit = sum(d['debit'] for d in details_data)
                total_credit = sum(d['credit'] for d in details_data)
                if total_debit != total_credit:
                    errors.append(
                        f'傳票 "{key}" （日期 {date}）借貸不平衡：'
                        f'借方 {total_debit}，貸方 {total_credit}'
                    )
                    continue

                # Determine voucher_no and status
                explicit_no = str(first_row[0] or '').strip()
                status_raw = str(first_row[3] or '草稿').strip()
                status = _STATUS_MAP.get(status_raw, Voucher.Status.DRAFT)
                desc = str(first_row[2] or '').strip()

                if not explicit_no:
                    # Auto-generate
                    date_str_fmt = date.strftime('%Y%m%d')
                    last = (
                        Voucher.objects
                        .filter(voucher_no__startswith=f'VOU-{date_str_fmt}-')
                        .order_by('-voucher_no')
                        .values_list('voucher_no', flat=True)
                        .first()
                    )
                    seq = int(last.split('-')[-1]) + 1 if last else 1
                    explicit_no = f'VOU-{date_str_fmt}-{seq:03d}'

                voucher = Voucher.objects.create(
                    voucher_no=explicit_no,
                    date=date,
                    description=desc,
                    status=status,
                    source=Voucher.Source.MANUAL,
                    created_by=request.user,
                )
                for d in details_data:
                    VoucherDetail.objects.create(voucher=voucher, **d)
                    detail_count += 1
                created_count += 1

        return {'created': created_count, 'details': detail_count, 'errors': errors}

    # ------------------------------------------------------------------ export

    def export_view(self, _request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '傳票資料'

        ws.append(_VOUCHER_HEADERS)
        hdr_font = Font(bold=True)
        hdr_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')

        vouchers = (
            Voucher.objects
            .filter(is_deleted=False)
            .prefetch_related('details__account')
            .order_by('date', 'voucher_no')
        )

        status_display = {Voucher.Status.DRAFT: '草稿', Voucher.Status.POSTED: '已過帳'}

        for v in vouchers:
            details = list(v.details.all())
            if details:
                for d in details:
                    ws.append([
                        v.voucher_no,
                        v.date.strftime('%Y-%m-%d'),
                        v.description,
                        status_display.get(v.status, v.status),
                        d.account.code,
                        int(d.debit) if d.debit else '',
                        int(d.credit) if d.credit else '',
                        d.company_id,
                        d.department,
                        d.project,
                        d.remark,
                    ])
            else:
                ws.append([
                    v.voucher_no,
                    v.date.strftime('%Y-%m-%d'),
                    v.description,
                    status_display.get(v.status, v.status),
                    '', '', '', '', '', '', '',
                ])

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # Reference sheet: account list
        ws2 = wb.create_sheet('科目代碼參考')
        ws2.append(['科目代碼', '科目名稱', '類別'])
        category_display = dict(Account.Category.choices)
        for acc in Account.objects.filter(is_active=True).order_by('code'):
            ws2.append([acc.code, acc.name, category_display.get(acc.category, acc.category)])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="voucher_export.xlsx"'
        wb.save(response)
        return response

    # ------------------------------------------------------------------ template

    def download_template(self, _request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '傳票匯入範本'
        ws.append(_VOUCHER_HEADERS)

        hdr_font = Font(bold=True)
        hdr_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')

        # Example: two-line voucher (debit 1110 / credit 400002)
        ws.append(['VOU-20250115-001', '2025-01-15', '記帳收入', '草稿', '1123', 5000, '', '', '', '', ''])
        ws.append(['VOU-20250115-001', '2025-01-15', '記帳收入', '草稿', '400002', '', 5000, '', '', '', ''])

        ws2 = wb.create_sheet('科目代碼參考')
        ws2.append(['科目代碼', '科目名稱', '類別'])
        category_display = dict(Account.Category.choices)
        for acc in Account.objects.filter(is_active=True).order_by('code'):
            ws2.append([acc.code, acc.name, category_display.get(acc.category, acc.category)])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="voucher_template.xlsx"'
        wb.save(response)
        return response

@admin.register(Receivable)
class ReceivableAdmin(admin.ModelAdmin):
    list_display = ('receivable_no', 'date', 'company_name', 'unified_business_no', 'assistant', 'is_deleted')
    list_filter = ('is_deleted',)
    search_fields = ('receivable_no', 'company_name', 'unified_business_no')
    actions = ['restore_receivables']
    list_per_page = 25

    def get_queryset(self, _request):
        return self.model._default_manager.all()

    @admin.action(description='還原選取的應收帳款（取消軟刪除）')
    def restore_receivables(self, request, queryset):
        updated = queryset.update(is_deleted=False)
        self.message_user(request, f'已成功還原 {updated} 筆應收帳款。')

@admin.register(VoucherDetail)
class VoucherDetailAdmin(admin.ModelAdmin):
    list_display = ('voucher', 'account', 'debit', 'credit', 'company_id', 'department', 'project')
    list_filter = ('account__category',)
    search_fields = ('voucher__voucher_no', 'account__name', 'remark')
    list_per_page = 25

@admin.register(FixedAsset)
class FixedAssetAdmin(admin.ModelAdmin):
    list_display = ('asset_no', 'name', 'purchase_date', 'cost', 'status')
    list_filter = ('status', 'purchase_date')
    search_fields = ('asset_no', 'name')
    list_per_page = 25

@admin.register(ReceivableNotification)
class ReceivableNotificationAdmin(admin.ModelAdmin):
    list_display = ('receivable', 'threshold_days', 'channel', 'success', 'sent_at', 'error_message')
    list_filter = ('channel', 'threshold_days', 'success')
    search_fields = ('receivable__receivable_no', 'receivable__company_name')
    readonly_fields = ('receivable', 'threshold_days', 'channel', 'sent_at', 'success', 'error_message')
    ordering = ('-sent_at',)
    list_per_page = 25

    def has_add_permission(self, _request):
        return False


@admin.register(Collection)
class CollectionAdmin(admin.ModelAdmin):
    list_display = ('collection_no', 'date', 'receivable', 'method', 'total', 'is_posted', 'is_deleted')
    list_filter = ('is_posted', 'method', 'is_deleted')
    search_fields = ('collection_no', 'receivable__company_name')
    actions = ['restore_collections']
    list_per_page = 25

    def get_queryset(self, _request):
        return self.model._default_manager.all()

    @admin.action(description='還原選取的收款紀錄（取消軟刪除）')
    def restore_collections(self, request, queryset):
        updated = queryset.update(is_deleted=False)
        self.message_user(request, f'已成功還原 {updated} 筆收款紀錄。')


@admin.register(PreCollection)
class PreCollectionAdmin(admin.ModelAdmin):
    list_display = (
        'pre_collection_no', 'date', 'company_name', 'unified_business_no',
        'amount', 'method', 'status', 'transaction_no', 'is_deleted',
    )
    list_filter = ('status', 'method', 'is_deleted', 'date')
    search_fields = ('pre_collection_no', 'company_name', 'unified_business_no', 'transaction_no')
    readonly_fields = (
        'pre_collection_no', 'source_content_type', 'source_id',
        'matched_receivable', 'matched_collection', 'created_at', 'updated_at',
    )
    ordering = ('-date', '-pre_collection_no')
    list_per_page = 25
    actions = ['restore_pre_collections']

    def get_queryset(self, _request):
        return self.model._default_manager.all()

    @admin.action(description='還原選取的預收款項（取消軟刪除）')
    def restore_pre_collections(self, request, queryset):
        updated = queryset.update(is_deleted=False)
        self.message_user(request, f'已成功還原 {updated} 筆預收款項。')

    def delete_model(self, _request, obj):
        obj.is_deleted = True
        obj.save(update_fields=['is_deleted', 'updated_at'])

    def delete_queryset(self, _request, queryset):
        queryset.update(is_deleted=True)
