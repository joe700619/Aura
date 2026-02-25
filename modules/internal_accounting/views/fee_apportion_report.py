from django.views.generic import TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum, Count, Q, F
from django.utils import timezone
from django.http import HttpResponse
from datetime import date

from ..models.receivable import ReceivableFeeApportion, Receivable
from modules.hr.models.employee import Employee


class FeeApportionReportView(LoginRequiredMixin, TemplateView):
    template_name = 'report/fee_apportion_report.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Date range defaults: current month
        today = date.today()
        date_from = self.request.GET.get('date_from', today.replace(day=1).isoformat())
        date_to = self.request.GET.get('date_to', today.isoformat())
        employee_id = self.request.GET.get('employee', '')
        q = self.request.GET.get('q', '')
        is_posted = self.request.GET.get('is_posted', 'all')

        # Base queryset
        qs = ReceivableFeeApportion.objects.select_related(
            'receivable', 'employee'
        ).filter(
            receivable__created_at__date__gte=date_from,
            receivable__created_at__date__lte=date_to,
            receivable__is_deleted=False,
        )

        # Filter by employee
        if employee_id:
            qs = qs.filter(employee_id=employee_id)

        # Search by customer name or receivable no
        if q:
            qs = qs.filter(
                Q(receivable__company_name__icontains=q) |
                Q(receivable__receivable_no__icontains=q) |
                Q(task_description__icontains=q)
            )

        # Filter by is_posted
        if is_posted == 'yes':
            qs = qs.filter(receivable__is_posted=True)
        elif is_posted == 'no':
            qs = qs.filter(receivable__is_posted=False)

        # Summary by employee
        summary_qs = qs.values(
            'employee__id',
            'employee__name',
            'employee__employee_number',
        ).annotate(
            total_amount=Sum('amount'),
            total_count=Count('id'),
        ).order_by('-total_amount')

        # Build summary with attached details
        summary_data = []
        for item in summary_qs:
            emp_id = item['employee__id']
            details = qs.filter(employee_id=emp_id).order_by('-receivable__created_at')
            item['details'] = details
            summary_data.append(item)

        # Grand total
        grand_total = qs.aggregate(
            total_amount=Sum('amount'),
            total_count=Count('id'),
        )

        # Employee list for filter dropdown
        employees = Employee.objects.filter(
            is_active=True
        ).order_by('name')

        context.update({
            'date_from': date_from,
            'date_to': date_to,
            'employee_id': employee_id,
            'q': q,
            'is_posted': is_posted,
            'summary_data': summary_data,
            'grand_total': grand_total,
            'employees': employees,
        })
        return context


class FeeApportionExportView(LoginRequiredMixin, View):
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        # Same filters as the report view
        today = date.today()
        date_from = request.GET.get('date_from', today.replace(day=1).isoformat())
        date_to = request.GET.get('date_to', today.isoformat())
        employee_id = request.GET.get('employee', '')
        q = request.GET.get('q', '')

        qs = ReceivableFeeApportion.objects.select_related(
            'receivable', 'employee'
        ).filter(
            receivable__created_at__date__gte=date_from,
            receivable__created_at__date__lte=date_to,
            receivable__is_deleted=False,
        )

        if employee_id:
            qs = qs.filter(employee_id=employee_id)
        if q:
            qs = qs.filter(
                Q(receivable__company_name__icontains=q) |
                Q(receivable__receivable_no__icontains=q) |
                Q(task_description__icontains=q)
            )

        qs = qs.order_by('employee__name', '-receivable__created_at')

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '公費分攤明細'

        # Header styling
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f'公費分攤明細報表 ({date_from} ~ {date_to})'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Headers
        headers = ['員工編號', '員工姓名', '應收單號', '客戶名稱', '執行項目', '分攤比例(%)', '分攤金額']
        ws.append([])  # Empty row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        row_num = 4
        total_amount = 0
        for item in qs:
            ws.cell(row=row_num, column=1, value=item.employee.employee_number if item.employee else '').border = thin_border
            ws.cell(row=row_num, column=2, value=item.employee.name if item.employee else '').border = thin_border
            ws.cell(row=row_num, column=3, value=item.receivable.receivable_no or '').border = thin_border
            ws.cell(row=row_num, column=4, value=item.receivable.company_name or '').border = thin_border
            ws.cell(row=row_num, column=5, value=item.task_description or '').border = thin_border
            cell_ratio = ws.cell(row=row_num, column=6, value=float(item.ratio))
            cell_ratio.border = thin_border
            cell_ratio.alignment = Alignment(horizontal='right')
            cell_amount = ws.cell(row=row_num, column=7, value=int(item.amount))
            cell_amount.border = thin_border
            cell_amount.number_format = '#,##0'
            cell_amount.alignment = Alignment(horizontal='right')
            total_amount += int(item.amount)
            row_num += 1

        # Total row
        ws.cell(row=row_num, column=6, value='合計').font = Font(bold=True)
        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal='right')
        ws.cell(row=row_num, column=6).border = thin_border
        total_cell = ws.cell(row=row_num, column=7, value=total_amount)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0'
        total_cell.alignment = Alignment(horizontal='right')
        total_cell.border = thin_border

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15

        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'fee_apportion_{date_from}_{date_to}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
