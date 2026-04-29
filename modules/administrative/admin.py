from decimal import Decimal, InvalidOperation
from collections import defaultdict

from django.contrib import admin
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import path
from import_export.admin import ImportExportModelAdmin
import openpyxl

from .models import TaxTemplate, TaxTaskInstance, ClientTaxEvent
from .models.document_receipt import DocumentReceipt
from .models.document_dispatch import DocumentDispatch, DocumentDispatchItem
from .models.seal_procurement import SealProcurement, SealProcurementItem
from .models.advance_payment import AdvancePayment, AdvancePaymentDetail


class DocumentDispatchItemInline(admin.TabularInline):
    model = DocumentDispatchItem
    extra = 0
    fields = ['customer', 'tax_id', 'address', 'postage', 'is_absorbed_by_customer', 'is_notified', 'is_deleted']
    readonly_fields = ['is_deleted']


@admin.register(DocumentDispatch)
class DocumentDispatchAdmin(admin.ModelAdmin):
    list_display = ['date', 'dispatch_method', 'is_deleted', 'created_at']
    list_filter = ['is_deleted', 'dispatch_method']
    date_hierarchy = 'date'
    inlines = [DocumentDispatchItemInline]


@admin.register(DocumentReceipt)
class DocumentReceiptAdmin(admin.ModelAdmin):
    list_display = ['receipt_date', 'customer', 'subject', 'category', 'status', 'is_deleted']
    list_filter = ['status', 'category', 'is_deleted']
    search_fields = ['subject', 'customer__name']
    date_hierarchy = 'receipt_date'

SEAL_TYPE_CODE_MAP = {label: code for code, label in SealProcurementItem.SEAL_TYPE_CHOICES}


@admin.register(SealProcurement)
class SealProcurementAdmin(admin.ModelAdmin):
    list_display = ['created_at', 'company_name', 'unified_business_no', 'seal_cost_subtotal', 'is_paid', 'is_deleted']
    list_filter = ['is_deleted', 'is_paid', 'transfer_to_inventory']
    search_fields = ['company_name', 'unified_business_no']
    date_hierarchy = 'created_at'
    change_list_template = 'admin/administrative/sealprocurement/change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import/', self.admin_site.admin_view(self.import_view), name='sealprocurement_import'),
            path('import/template/', self.admin_site.admin_view(self.download_template), name='sealprocurement_template'),
            path('export/', self.admin_site.admin_view(self.export_view), name='sealprocurement_export'),
        ]
        return custom_urls + urls

    def import_view(self, request):
        if request.method == 'POST':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                self.message_user(request, '請選擇 Excel 檔案', level='error')
                return redirect('.')

            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                ws = wb.active

                rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if any(row)]

                # Group by (統一編號, 採購日期) → one SealProcurement per group
                groups = defaultdict(list)
                for row in rows:
                    ubn = str(row[0] or '').strip()
                    date_val = row[8]
                    if hasattr(date_val, 'strftime'):
                        date_key = date_val.strftime('%Y-%m-%d')
                    else:
                        date_key = str(date_val or '').strip()
                    groups[(ubn, date_key)].append(row)

                created_count = 0
                item_count = 0

                for (ubn, date_key), items in groups.items():
                    first = items[0]

                    procurement = SealProcurement(
                        unified_business_no=ubn,
                        company_name=str(first[1] or '').strip(),
                        main_contact=str(first[2] or '').strip(),
                        mobile=str(first[3] or '').strip(),
                        phone=str(first[4] or '').strip(),
                        address=str(first[5] or '').strip(),
                        line_id=str(first[6] or '').strip(),
                        room_id=str(first[7] or '').strip(),
                        transfer_to_inventory=True,
                        note=str(first[14] or '').strip(),
                    )
                    procurement.save()

                    # Override created_at (auto_now_add bypassed via update)
                    if date_key:
                        try:
                            from django.utils import timezone
                            from datetime import datetime
                            dt = datetime.strptime(date_key, '%Y-%m-%d')
                            aware_dt = timezone.make_aware(dt)
                            SealProcurement.objects.filter(pk=procurement.pk).update(created_at=aware_dt)
                        except ValueError:
                            pass

                    created_count += 1

                    for row in items:
                        raw_seal = str(row[9] or '').strip()
                        # Accept both code (large_self) and label (大章(自留))
                        seal_type = SEAL_TYPE_CODE_MAP.get(raw_seal, raw_seal)
                        try:
                            quantity = int(row[10] or 1)
                        except (ValueError, TypeError):
                            quantity = 1
                        try:
                            unit_price = Decimal(str(row[12] or 0))
                        except InvalidOperation:
                            unit_price = Decimal(0)
                        absorbed_raw = str(row[13] or '').strip()
                        is_absorbed = absorbed_raw in ('是', 'Y', 'y', 'True', 'true', '1')

                        SealProcurementItem.objects.create(
                            procurement=procurement,
                            movement_type='purchase',
                            seal_type=seal_type,
                            quantity=quantity,
                            name_or_address=str(row[11] or '').strip(),
                            unit_price=unit_price,
                            is_absorbed_by_customer=is_absorbed,
                        )
                        item_count += 1

                    procurement.recalculate_subtotal()

                self.message_user(request, f'成功匯入 {created_count} 張採購單，共 {item_count} 筆明細')
                return redirect('../')

            except Exception as e:
                self.message_user(request, f'匯入失敗：{e}', level='error')
                return redirect('.')

        context = self.admin_site.each_context(request)
        context['title'] = '匯入印章採購資料'
        context['opts'] = self.model._meta
        context['seal_type_choices'] = SealProcurementItem.SEAL_TYPE_CHOICES
        context['column_info'] = [
            {'name': '統一編號', 'note': '公司統一編號，與採購日期合併決定一張採購單', 'example': '12345678'},
            {'name': '公司名稱', 'note': '', 'example': '測試公司有限公司'},
            {'name': '主要聯絡人', 'note': '', 'example': '王小明'},
            {'name': '手機', 'note': '', 'example': '0912345678'},
            {'name': '電話', 'note': '', 'example': '02-12345678'},
            {'name': '通訊地址', 'note': '', 'example': '台北市中正區...'},
            {'name': 'Line ID', 'note': '可留空', 'example': ''},
            {'name': 'Room ID', 'note': '可留空', 'example': ''},
            {'name': '採購日期', 'note': '格式 YYYY-MM-DD', 'example': '2025-01-15'},
            {'name': '印章種類', 'note': '填入代碼，見下表', 'example': 'large_self'},
            {'name': '數量', 'note': '整數', 'example': '2'},
            {'name': '名稱/地址', 'note': '印章上的文字，可留空', 'example': '公司大章'},
            {'name': '單價', 'note': '數字', 'example': '150'},
            {'name': '客戶吸收', 'note': '是 / 否', 'example': '否'},
            {'name': '備註', 'note': '可留空，同一採購單只取第一列', 'example': ''},
        ]
        return render(request, 'admin/administrative/sealprocurement/import.html', context)

    def export_view(self, _request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '印章採購資料'

        headers = [
            '統一編號', '公司名稱', '主要聯絡人', '手機', '電話', '通訊地址',
            'Line ID', 'Room ID', '採購日期(YYYY-MM-DD)', '印章種類',
            '數量', '名稱/地址', '單價', '客戶吸收(是/否)', '備註',
        ]
        ws.append(headers)

        # Bold header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        procurements = (
            SealProcurement.objects
            .filter(is_deleted=False)
            .prefetch_related('items')
            .order_by('created_at', 'company_name')
        )

        for proc in procurements:
            items = [i for i in proc.items.all() if not i.is_deleted]
            if items:
                for item in items:
                    ws.append([
                        proc.unified_business_no,
                        proc.company_name,
                        proc.main_contact,
                        proc.mobile,
                        proc.phone,
                        proc.address,
                        proc.line_id,
                        proc.room_id,
                        proc.created_at.strftime('%Y-%m-%d') if proc.created_at else '',
                        item.seal_type,
                        item.quantity,
                        item.name_or_address,
                        int(item.unit_price),
                        '是' if item.is_absorbed_by_customer else '否',
                        proc.note,
                    ])
            else:
                # Procurement with no items — export header row only
                ws.append([
                    proc.unified_business_no,
                    proc.company_name,
                    proc.main_contact,
                    proc.mobile,
                    proc.phone,
                    proc.address,
                    proc.line_id,
                    proc.room_id,
                    proc.created_at.strftime('%Y-%m-%d') if proc.created_at else '',
                    '', '', '', '', '', proc.note,
                ])

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        ws2 = wb.create_sheet('印章種類說明')
        ws2.append(['代碼', '名稱'])
        for code, label in SealProcurementItem.SEAL_TYPE_CHOICES:
            ws2.append([code, label])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="seal_procurement_export.xlsx"'
        wb.save(response)
        return response

    def download_template(self, _request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '印章採購匯入範本'
        headers = [
            '統一編號', '公司名稱', '主要聯絡人', '手機', '電話', '通訊地址',
            'Line ID', 'Room ID', '採購日期(YYYY-MM-DD)', '印章種類',
            '數量', '名稱/地址', '單價', '客戶吸收(是/否)', '備註',
        ]
        ws.append(headers)
        ws.append([
            '12345678', '測試公司有限公司', '王小明', '0912345678', '02-12345678',
            '台北市中正區重慶南路一段1號', '', '', '2025-01-15',
            'large_self', 2, '公司大章', 150, '否', '',
        ])
        ws.append([
            '12345678', '測試公司有限公司', '王小明', '0912345678', '02-12345678',
            '台北市中正區重慶南路一段1號', '', '', '2025-01-15',
            'small_self', 1, '公司小章', 50, '否', '',
        ])

        ws2 = wb.create_sheet('印章種類說明')
        ws2.append(['代碼(填入主表)', '顯示名稱'])
        for code, label in SealProcurementItem.SEAL_TYPE_CHOICES:
            ws2.append([code, label])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="seal_procurement_template.xlsx"'
        wb.save(response)
        return response

@admin.register(SealProcurementItem)
class SealProcurementItemAdmin(admin.ModelAdmin):
    list_display = ['procurement', 'movement_type', 'seal_type', 'quantity', 'unit_price', 'subtotal', 'is_deleted']
    list_filter = ['is_deleted', 'movement_type', 'seal_type']
    search_fields = ['procurement__company_name', 'name_or_address']

class AdvancePaymentDetailInline(admin.TabularInline):
    model = AdvancePaymentDetail
    extra = 0
    fields = ['is_customer_absorbed', 'customer', 'unified_business_no', 'reason', 'amount', 'payment_type', 'is_deleted']
    readonly_fields = ['is_deleted']

@admin.register(AdvancePayment)
class AdvancePaymentAdmin(admin.ModelAdmin):
    list_display = ['advance_no', 'date', 'applicant', 'total_amount', 'is_posted', 'is_deleted', 'created_at']
    list_filter = ['is_deleted', 'is_posted', 'date']
    search_fields = ['advance_no', 'description']
    date_hierarchy = 'date'
    readonly_fields = ['advance_no', 'created_at', 'updated_at']
    inlines = [AdvancePaymentDetailInline]

@admin.register(AdvancePaymentDetail)
class AdvancePaymentDetailAdmin(admin.ModelAdmin):
    list_display = ['advance_payment', 'reason', 'amount', 'is_customer_absorbed', 'payment_type', 'is_deleted']
    list_filter = ['is_deleted', 'is_customer_absorbed', 'payment_type']
    search_fields = ['advance_payment__advance_no', 'reason', 'unified_business_no']

@admin.register(TaxTemplate)
class TaxTemplateAdmin(ImportExportModelAdmin):
    list_display = ['name', 'is_recurring', 'recurring_months', 'deadline_day', 'source_type']
    list_filter = ['is_recurring', 'source_type']
    search_fields = ['name']

@admin.register(TaxTaskInstance)
class TaxTaskInstanceAdmin(ImportExportModelAdmin):
    list_display = ['title', 'template', 'year', 'month', 'deadline', 'is_completed', 'completed_clients', 'total_clients']
    list_filter = ['year', 'month', 'is_completed', 'template']
    search_fields = ['title']
    date_hierarchy = 'deadline'


@admin.register(ClientTaxEvent)
class ClientTaxEventAdmin(ImportExportModelAdmin):
    list_display = ['title', 'deadline', 'urgent_days', 'is_active', 'sort_order']
    list_filter = ['is_active']
    search_fields = ['title']
    list_editable = ['is_active', 'sort_order']
    date_hierarchy = 'deadline'
