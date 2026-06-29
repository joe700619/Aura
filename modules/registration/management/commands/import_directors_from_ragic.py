"""
從 Ragic 匯出的「董監事名單」檔（.xlsx / .csv）匯入：
  - ShareholderRegister（用統編 get_or_create）
  - DirectorSupervisor（該公司底下的董監事，子表）
  - Shareholder 主檔（放大鏡搜尋來源，可用 --no-shareholders 關閉）

設計重點（對應「資料可能有錯、想先試跑」的需求）：
  * 預設 dry-run：只檢查、逐列回報問題，完全不寫入。確認無誤後加 --commit 才真寫。
  * 整批包在 transaction：--commit 時全成功才寫，中途出錯整批回滾。
  * 可重跑（idempotent）：register 用統編、director 用 (register, 姓名, 身分證)、
    shareholder 用身分證 來 get_or_create / update_or_create，重跑不會產生重複。
  * 有問題的列會被略過、其餘照常匯入；修好資料後可直接重跑補上。

用法（在 Docker 容器內）：
  docker compose exec web python manage.py import_directors_from_ragic 檔名.xlsx          # 試跑
  docker compose exec web python manage.py import_directors_from_ragic 檔名.xlsx --commit  # 實際匯入
"""
import csv
from datetime import date, datetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from modules.registration.models import (
    ShareholderRegister, DirectorSupervisor, Shareholder,
)

# 身分別（Ragic 文字）→ DirectorSupervisor.Title
TITLE_MAP = {
    '董事長': DirectorSupervisor.Title.CHAIRMAN,
    '董事':   DirectorSupervisor.Title.DIRECTOR,
    '監察人': DirectorSupervisor.Title.SUPERVISOR,
    '經理人': DirectorSupervisor.Title.MANAGER,
}

# 國籍代碼 → 顯示文字（DirectorSupervisor.nationality 是自由字串）
NATIONALITY_DISPLAY = {'TW': '中華民國', 'CN': '中國大陸', 'HK': '香港', 'KR': '韓國'}
# Shareholder.nationality 是 choices，存代碼
SH_NATIONALITY_CODES = {c[0] for c in Shareholder.Nationality.choices}

# Ragic 表頭 → 內部 key（_norm 會把「身份」統一成「身分」、去除前後空白）
COLUMNS = {
    '股東姓名': 'name',
    '身分證字號': 'id_number',
    '國籍': 'nationality',
    '生日': 'birth_date',
    '身分別': 'title',
    '持有股份': 'shares_held',
    '統一編號': 'ubn',
    '公司名稱': 'company_name',
    '所代表法人': 'entity_name',
    '代表法人統一編號': 'entity_no',
    '代表法人統編': 'entity_no',  # 別名
    # 金額、公司類別 目前沒有對應的 model 欄位，故不列入（略過）
}


def _norm(s):
    return (s or '').strip().replace('身份', '身分')


def _text(v):
    return str(v).strip() if v is not None else ''


def parse_ubn(v):
    """統編：Excel 數字會掉開頭的 0，補回 8 碼。"""
    s = _text(v)
    if s.endswith('.0'):
        s = s[:-2]
    s = ''.join(ch for ch in s if ch.isdigit())
    return s.zfill(8) if s else ''


def parse_int(v):
    """持有股份：去千分位、'-'/空白視為 0。"""
    s = _text(v).replace(',', '').replace('，', '')
    if s in ('', '-', '－'):
        return 0
    if s.endswith('.0'):
        s = s[:-2]
    return int(s)  # 失敗會丟 ValueError，由呼叫端收集


def parse_date(v):
    if v in (None, '', '-', '－'):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _text(v)
    for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%Y.%m.%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f'無法解析生日 {s!r}')


def read_rows(path):
    if path.lower().endswith('.csv'):
        with open(path, encoding='utf-8-sig', newline='') as f:
            return list(csv.reader(f))
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    return [list(r) for r in ws.iter_rows(values_only=True)]


class Command(BaseCommand):
    help = '從 Ragic 匯出的董監事名單檔（xlsx/csv）匯入 register + 董監事（預設 dry-run）'

    def add_arguments(self, parser):
        parser.add_argument('file', help='Ragic 匯出的 .xlsx 或 .csv 路徑')
        parser.add_argument('--commit', action='store_true',
                            help='實際寫入；不加則為 dry-run，只檢查不寫入')
        parser.add_argument('--no-shareholders', action='store_true',
                            help='不要同步 Shareholder 主檔（放大鏡搜尋來源）')
        parser.add_argument('--update-shareholders', action='store_true',
                            help='允許用 Ragic 覆蓋既有股東主檔的姓名/國籍/生日（預設只新增、不覆蓋）')

    def handle(self, *args, **opts):
        path = opts['file']
        commit = opts['commit']
        sync_sh = not opts['no_shareholders']
        update_sh = opts['update_shareholders']

        rows = read_rows(path)
        if not rows:
            raise CommandError('檔案是空的')

        # 表頭對應
        idx = {}
        for i, h in enumerate(rows[0]):
            key = COLUMNS.get(_norm(h))
            if key:
                idx[key] = i
        missing = [c for c in ('name', 'title', 'ubn', 'company_name') if c not in idx]
        if missing:
            raise CommandError(f'缺少必要欄位 {missing}；實際表頭為：{rows[0]}')

        def cell(row, key):
            i = idx.get(key)
            return row[i] if i is not None and i < len(row) else None

        parsed, errors = [], []
        for n, row in enumerate(rows[1:], start=2):  # 2 = 扣掉表頭後的真實列號
            if row is None or all(c in (None, '') for c in row):
                continue
            name = _text(cell(row, 'name'))
            ubn = parse_ubn(cell(row, 'ubn'))
            company = _text(cell(row, 'company_name'))
            title_raw = _text(cell(row, 'title'))
            row_errs = []

            if not name:
                row_errs.append('缺姓名')
            if not ubn:
                row_errs.append('統編缺漏或非數字')
            if not company:
                row_errs.append('缺公司名稱')
            title = TITLE_MAP.get(title_raw)
            if title is None:
                row_errs.append(f'未知身分別「{title_raw}」')
            try:
                birth = parse_date(cell(row, 'birth_date'))
            except ValueError as e:
                birth, _ = None, row_errs.append(str(e))
            try:
                shares = parse_int(cell(row, 'shares_held'))
            except ValueError:
                shares = 0
                row_errs.append('持有股份非數字')

            if row_errs:
                errors.append((n, name or '(無名)', row_errs))
                continue

            parsed.append({
                'row': n, 'ubn': ubn, 'company': company, 'title': title,
                'name': name, 'id_number': _text(cell(row, 'id_number')),
                'nat': _text(cell(row, 'nationality')), 'birth': birth, 'shares': shares,
                'entity_name': _text(cell(row, 'entity_name')),
                'entity_no': parse_ubn(cell(row, 'entity_no')),  # 法人統編同樣補回開頭 0
            })

        # ── 報告 ──
        self.stdout.write(f'總資料列 {len(rows) - 1} 筆 → 可匯入 {len(parsed)}、有問題 {len(errors)}')
        for n, name, errs in errors:
            self.stdout.write(self.style.WARNING(f'  第 {n} 列 {name}：{"；".join(errs)}'))

        if not commit:
            self.stdout.write(self.style.NOTICE(
                '\n[DRY-RUN] 未寫入任何資料。確認上面沒問題後，加 --commit 實際匯入。'))
            return

        if errors:
            self.stdout.write(self.style.WARNING('\n有問題的列會略過，其餘照常匯入。'))

        # ── 實際寫入（整批 transaction）──
        n_reg = n_dir = n_sh = 0
        reg_cache, order_next = {}, {}
        with transaction.atomic():
            for p in parsed:
                reg = reg_cache.get(p['ubn'])
                if reg is None:
                    reg, created = ShareholderRegister.objects.get_or_create(
                        unified_business_no=p['ubn'],
                        defaults={'company_name': p['company']})
                    reg_cache[p['ubn']] = reg
                    order_next[p['ubn']] = reg.directors.count()
                    n_reg += int(created)

                _, dcreated = DirectorSupervisor.objects.get_or_create(
                    register=reg, name=p['name'], id_number=p['id_number'],
                    defaults={
                        'title': p['title'],
                        'nationality': NATIONALITY_DISPLAY.get(p['nat'], p['nat']),
                        'birth_date': p['birth'],
                        'shares_held': p['shares'],
                        'entity_name': p['entity_name'],
                        'entity_no': p['entity_no'],
                        'order': order_next[p['ubn']],
                    })
                if dcreated:
                    order_next[p['ubn']] += 1
                    n_dir += 1

                if sync_sh and p['id_number']:
                    sh_nat = p['nat'] if p['nat'] in SH_NATIONALITY_CODES else Shareholder.Nationality.TW
                    defaults = {'name': p['name'], 'nationality': sh_nat, 'birthday': p['birth']}
                    if update_sh:
                        # 允許覆蓋既有主檔
                        _, sh_created = Shareholder.objects.update_or_create(
                            id_number=p['id_number'], defaults=defaults)
                    else:
                        # 預設：只新增、絕不動既有主檔
                        _, sh_created = Shareholder.objects.get_or_create(
                            id_number=p['id_number'], defaults=defaults)
                    n_sh += int(sh_created)

        self.stdout.write(self.style.SUCCESS(
            f'\n完成：新增 register {n_reg}、董監事 {n_dir}、新增股東主檔 {n_sh} '
            f'（{"允許覆蓋既有主檔" if update_sh else "既有主檔未更動"}；可重跑、不會重複）'))
